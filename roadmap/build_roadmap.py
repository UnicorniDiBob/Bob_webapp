#!/usr/bin/env python3
"""Genera BOB_Roadmap_Gantt.xlsx dal file roadmap.csv (source of truth).
Le barre del Gantt si disegnano da sole (conditional formatting) da
Status + Start + End; il mese corrente è evidenziato. La colonna Track
instrada le attività (Client/Pro = André, Internal = Lucio, Shared).

Uso:  python3 build_roadmap.py         # legge roadmap.csv, scrive l'xlsx
"""
import csv, os, sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
CSV = os.path.join(HERE, "roadmap.csv")
OUT = os.path.join(HERE, "BOB_Roadmap_Gantt.xlsx")

NAVY="1F3864"; SECT="D6E4F0"; ZEBRA="F2F6FB"; WHITE="FFFFFF"; INK="1A1A2E"
GREEN="2E7D32"; BLUE="1976D2"; GREY="9E9E9E"; AMBER="FFC000"; REDHDR="C00000"
TRACK_FILL={"Client/Pro":"E6E9F7","Internal":"DCF1EA","Shared":"ECECE8"}
TRACK_TEXT={"Client/Pro":"3730A3","Internal":"0F6E56","Shared":"5F5E5A"}

def D(s): return datetime.strptime(s,"%Y-%m-%d") if s else ""

def main():
    with open(CSV,encoding="utf-8") as f:
        R=list(csv.DictReader(f))

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Roadmap"
    months=[]; y,m=2026,6
    while (y,m)<=(2027,12):
        months.append(datetime(y,m,1)); m+=1
        if m>12: m=1;y+=1
    NMON=len(months); MON0=9; LASTCOL=MON0+NMON-1
    thin=Side(style="thin",color="D9D9D9"); border=Border(thin,thin,thin,thin)

    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=LASTCOL)
    t=ws.cell(1,1,"BOB — Project Roadmap & Gantt   |   Jun 2026 → Dec 2027   |   Tracker: stato, date, completamento, Track (owner)")
    t.font=Font(name="Arial",bold=True,size=13,color=WHITE); t.fill=PatternFill("solid",fgColor=NAVY)
    t.alignment=Alignment(vertical="center",indent=1); ws.row_dimensions[1].height=26

    heads=["#","Track","Owner","Task / Milestone","Status","Start","End","Done on"]
    for i,h in enumerate(heads,1):
        c=ws.cell(2,i,h); c.font=Font(name="Arial",bold=True,size=9,color=WHITE)
        c.fill=PatternFill("solid",fgColor=NAVY)
        c.alignment=Alignment(vertical="center",horizontal="left" if i==4 else "center",indent=1 if i==4 else 0); c.border=border
    for j,dt in enumerate(months):
        c=ws.cell(2,MON0+j,dt); c.number_format="mmm\\ yy"
        c.font=Font(name="Arial",bold=True,size=8,color=WHITE); c.fill=PatternFill("solid",fgColor=NAVY)
        c.alignment=Alignment(horizontal="center",vertical="bottom",textRotation=90); c.border=border
    ws.row_dimensions[2].height=44

    for col,w in zip("ABCDEFGH",[6,11,9,60,13,11,11,11]): ws.column_dimensions[col].width=w
    for j in range(NMON): ws.column_dimensions[get_column_letter(MON0+j)].width=4.6
    ws.freeze_panes="I3"

    r=3; data_start=3
    for row in R:
        if row["kind"]=="section":
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
            c=ws.cell(r,1,row["task"]); c.font=Font(name="Arial",bold=True,size=9,color=NAVY)
            c.fill=PatternFill("solid",fgColor=SECT); c.alignment=Alignment(vertical="center",indent=1)
            for ci in range(1,LASTCOL+1):
                cc=ws.cell(r,ci); cc.border=border
                if ci>8: cc.fill=PatternFill("solid",fgColor=SECT)
            ws.row_dimensions[r].height=18; r+=1; continue
        isproj=row["kind"]=="project"; tr=row["track"]
        zeb=PatternFill("solid",fgColor=ZEBRA) if not isproj else PatternFill("solid",fgColor=WHITE)
        vals=[row["id"],tr,row["owner"],row["task"],row["status"],D(row["start"]),D(row["end"]),D(row["done_on"])]
        for ci,v in enumerate(vals,1):
            c=ws.cell(r,ci,v); c.border=border
            c.font=Font(name="Arial",size=9,bold=isproj and ci==4,color=INK)
            c.fill=zeb if ci not in (2,) else PatternFill("solid",fgColor=TRACK_FILL.get(tr,WHITE))
            if ci==2 and tr:
                c.font=Font(name="Arial",size=8,bold=True,color=TRACK_TEXT.get(tr,INK)); c.alignment=Alignment(horizontal="center",vertical="center")
            elif ci in (6,7,8): c.number_format="dd\\ mmm\\ yy"; c.alignment=Alignment(horizontal="center")
            elif ci in (1,3,5): c.alignment=Alignment(horizontal="center")
            else: c.alignment=Alignment(vertical="center",indent=0 if isproj else 1,wrap_text=(ci==4))
        for j in range(NMON): ws.cell(r,MON0+j).border=border
        if not isproj: ws.row_dimensions[r].outlineLevel=1
        r+=1
    data_end=r-1

    rng=f"{get_column_letter(MON0)}{data_start}:{get_column_letter(LASTCOL)}{data_end}"
    mcol=get_column_letter(MON0)
    def rule(status,color):
        f=f'AND($F{data_start}<>"",{mcol}$2<=$G{data_start},EOMONTH({mcol}$2,0)>=$F{data_start},$E{data_start}="{status}")'
        return FormulaRule(formula=[f],fill=PatternFill("solid",fgColor=color))
    for st,col in [("Done",GREEN),("In progress",BLUE),("Planned",GREY),("Milestone",AMBER)]:
        ws.conditional_formatting.add(rng,rule(st,col))
    hdr=f"{mcol}2:{get_column_letter(LASTCOL)}2"
    ws.conditional_formatting.add(hdr,FormulaRule(formula=[f'AND({mcol}$2<=TODAY(),EOMONTH({mcol}$2,0)>=TODAY())'],fill=PatternFill("solid",fgColor=REDHDR)))

    lr=data_end+2
    ws.cell(lr,2,"LEGENDA").font=Font(name="Arial",bold=True,size=9,color=NAVY)
    leg=[("Done (verde)",GREEN),("In progress (blu)",BLUE),("Planned (grigio)",GREY),("Milestone / gate (ambra)",AMBER)]
    for i,(lab,col) in enumerate(leg):
        ws.cell(lr+1+i,2).fill=PatternFill("solid",fgColor=col); ws.cell(lr+1+i,4,lab).font=Font(name="Arial",size=9,color=INK)
    trk=[("Track: Client/Pro → André  (esperienza cliente e professionista)","Client/Pro"),
         ("Track: Internal → Lucio  (dati, dashboard, admin, privacy)","Internal"),
         ("Track: Shared  (infra, legale, GTM, milestone)","Shared")]
    for i,(lab,k) in enumerate(trk):
        ws.cell(lr+1+i,6).fill=PatternFill("solid",fgColor=TRACK_FILL[k]); ws.cell(lr+1+i,7,lab).font=Font(name="Arial",size=9,color=TRACK_TEXT[k])
        ws.merge_cells(start_row=lr+1+i,start_column=7,end_row=lr+1+i,end_column=13)
    note=("Fonte dati: roadmap.csv (modificalo e rilancia build_roadmap.py). "
          "Le barre si disegnano da Status+Start+End; 'Done on' = data di completamento effettiva. "
          "Mese corrente evidenziato in rosso. Email notifiche: pronte ma DORMIENTI fino al task 15.4.")
    ws.cell(lr+6,4,note).font=Font(name="Arial",size=8,italic=True,color="555555")
    ws.cell(lr+6,4).alignment=Alignment(wrap_text=True,vertical="top")
    ws.merge_cells(start_row=lr+6,start_column=4,end_row=lr+9,end_column=13)

    ws.sheet_view.showGridLines=False
    wb.save(OUT)
    done=sum(1 for x in R if x["status"]=="Done")
    print(f"OK: {OUT}  ({len(R)} rows, {done} done)")

if __name__=="__main__":
    main()
